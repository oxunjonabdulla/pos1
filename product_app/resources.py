from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget

from .models import Maxsulot, OrderItems, CartItems


class MaxsulotResource(resources.ModelResource):
    class Meta:
        model = Maxsulot
        fields = ('nomi', 'rasm', 'foydalanuvchi__username', 'razmer', 'qoshimcha')


class OrderItemsResource(resources.ModelResource):
    maxsulot = fields.Field(
        column_name='maxsulot',
        attribute='maxsulot',
        widget=ForeignKeyWidget(Maxsulot, 'nomi')
    )

    class Meta:
        model = OrderItems
        fields = ('maxsulot', 'soni', 'foydalanuvchi__username')


class CartItemsResource(resources.ModelResource):
    maxsulot = fields.Field(
        column_name='maxsulot',
        attribute='maxsulot',
        widget=ForeignKeyWidget(Maxsulot, 'nomi')
    )

    class Meta:
        model = CartItems
        fields = ('maxsulot', 'soni', 'foydalanuvchi__username')



from import_export import resources
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment



class OrderResource(resources.ModelResource):
    def export_to_template_xlsx(self, queryset, template_path):
        """
        Export data to an XLSX file using a predefined template.
        All products in an order will be concatenated in one row, and the row height will adjust automatically.
        """
        # Load the template
        try:
            wb = load_workbook(template_path)
        except FileNotFoundError:
            raise FileNotFoundError(f"Template file not found: {template_path}")

        ws = wb.active

        row_num = 2  # Starting row for data

        # Populate the data into the template
        for obj in queryset:
            products = obj.maxsulotlar.all()

            # Concatenate product details into a single string
            product_details = "\n".join([
                f"{product.maxsulot.nomi} ({product.maxsulot.razmer}) x {product.soni}"
                for product in products
            ]) if products.exists() else "No Products"

            ws[f"A{row_num}"] = obj.id  # Order ID
            ws[f"B{row_num}"] = product_details  # Products
            ws[f"C{row_num}"] = obj.foydalanuvchi.username  # User
            ws[f"D{row_num}"] = obj.jami_maxsulot  # Total Products

            # Format the Status field using choices
            ws[f"E{row_num}"] = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)

            # ws[f"F{row_num}"] = obj.qoshimcha_matn2  # Additional Text

            # Format the To field using choices
            ws[f"F{row_num}"] = dict(obj._meta.get_field('kimga').choices).get(obj.kimga, obj.kimga)

            # Format the Time field
            ws[f"G{row_num}"] = obj.created_at.strftime("%Y-%m-%d %H:%M:%S") if obj.created_at else ""

            # Adjust row height to fit product details
            row_height = max(len(product_details.split("\n")) * 15, 15)  # 15 points per line as a base
            ws.row_dimensions[row_num].height = row_height

            row_num += 1

        # Set uniform alignment for all cells
        for row in ws.iter_rows(min_row=2, max_row=row_num - 1):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Save to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output